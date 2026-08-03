# -*- coding: utf-8 -*-
"""
Dopisuje aktywo cash + reguły zakupu do a_config.xlsx (arkusze roi_def / roi_rules)
z zachowaniem Excel Table.

Użycie:
  cd app
  uv run python maintenance/add_cash_to_analyse_assets_config.py
  uv run python maintenance/add_cash_to_analyse_assets_config.py C:/sciezka/a_config.xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from analyse_assets.config_model import (
    CATALOG_SHEET,
    A_CONFIG_FILE_NAME,
    RULES_SHEET,
    AnalyseAssetsCatalog,
    AnalyseAssetsRules,
)
from importers.assets.pool_id import MBANK_EUR
from app_proc.data_root import get_online_data_root
from maintenance.export_analyse_assets_config import _catalog, _rules

_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")


def _parse_cell(cell_ref: str) -> tuple[str, int]:
    match = _CELL_REF.match(cell_ref.upper())
    if not match:
        raise ValueError(f"Niepoprawna referencja komórki: {cell_ref!r}")
    return match.group(1), int(match.group(2))


def _col_index(col_letters: str) -> int:
    value = 0
    for char in col_letters:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def _header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            headers[str(value).strip()] = col_idx
        if headers:
            return row_idx, headers
    raise ValueError(f"Brak wiersza nagłówków w arkuszu {ws.title!r}")


def _extend_table_rows(ws, header_row: int, last_data_row: int) -> None:
    for table in ws.tables.values():
        start_cell, end_cell = table.ref.split(":")
        start_col_letters, start_row = _parse_cell(start_cell)
        end_col_letters, end_row = _parse_cell(end_cell)
        if start_row != header_row:
            continue
        start_col = _col_index(start_col_letters)
        end_col = _col_index(end_col_letters)
        if last_data_row <= end_row:
            continue
        table.ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{last_data_row}"
        )


def _asset_ids(ws, header_row: int, headers: dict[str, int]) -> set[str]:
    col = headers[AnalyseAssetsCatalog.ASSET_ID]
    result: set[str] = set()
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        value = ws.cell(row_idx, col).value
        if value is None or str(value).strip() == "":
            continue
        result.add(str(value).strip())
    return result


def _append_row(ws, header_row: int, headers: dict[str, int], row: dict) -> int:
    new_row = (ws.max_row or header_row) + 1
    for key, value in row.items():
        if key not in headers:
            continue
        ws.cell(new_row, headers[key]).value = value
    return new_row


def add_cash(target: Path) -> None:
    wb = load_workbook(target)
    catalog_ws = wb[CATALOG_SHEET]
    rules_ws = wb[RULES_SHEET]

    cat_header_row, cat_headers = _header_row_and_columns(catalog_ws)
    existing = _asset_ids(catalog_ws, cat_header_row, cat_headers)

    cash_catalog = _catalog()
    cash_catalog = cash_catalog[cash_catalog[AnalyseAssetsCatalog.ASSET_ID] == "cash"]
    if cash_catalog.empty:
        raise RuntimeError("Brak wiersza cash w seedzie eksportera")

    last_cat_row = catalog_ws.max_row or cat_header_row
    if "cash" not in existing:
        row = cash_catalog.iloc[0].to_dict()
        last_cat_row = _append_row(catalog_ws, cat_header_row, cat_headers, row)
        print(f"Dodano katalog cash (pool_id={MBANK_EUR})")
    else:
        print("Katalog cash juz istnieje — pomijam")

    _extend_table_rows(catalog_ws, cat_header_row, last_cat_row)

    rules_header_row, rules_headers = _header_row_and_columns(rules_ws)
    rule_assets = _asset_ids(rules_ws, rules_header_row, rules_headers)
    last_rules_row = rules_ws.max_row or rules_header_row

    if "cash" not in rule_assets:
        cash_rules = _rules()
        cash_rules = cash_rules[cash_rules[AnalyseAssetsRules.ASSET_ID] == "cash"]
        for _, rule in cash_rules.iterrows():
            last_rules_row = _append_row(
                rules_ws,
                rules_header_row,
                rules_headers,
                rule.to_dict(),
            )
        print(f"Dodano {len(cash_rules)} regul zakupu cash")
    else:
        print("Reguly cash juz istnieja — pomijam")

    _extend_table_rows(rules_ws, rules_header_row, last_rules_row)
    wb.save(target)
    print(f"Zapisano: {target.resolve()}")


def main() -> None:
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else get_online_data_root() / A_CONFIG_FILE_NAME
    assert target.is_file(), target
    add_cash(target)


if __name__ == "__main__":
    main()
