# -*- coding: utf-8 -*-
"""
Migracja rocky-iv do procedury ROI (jak cash):
  - wiersze z arkusza rocky-iv → asset-evaluation (id=rocky-iv)
  - RODZAJ* w assets → assets.cash
  - usunięcie arkusza rocky-iv

Użycie:
  cd app
  uv run python maintenance/migrate_rocky_iv_to_roi.py
  uv run python maintenance/migrate_rocky_iv_to_roi.py --dry-run
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from importers.assets.data_model import (
    ASSET_EVALUATION_SHEET,
    AssetsDef,
    OperationDomain,
    Properties,
)
from importers.assets.read_assets import (
    ASSETS_FILE_NAME,
    get_assets_file,
)

ASSET_ID = "rocky-iv"
LEGACY_SHEET = "rocky-iv"
ASSETS_SHEET = "assets"
TARGET_KIND = "assets.cash"


def _header_map(ws, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, (ws.max_column or 0) + 1):
        value = ws.cell(header_row, col_idx).value
        if value is None:
            continue
        headers[str(value).strip()] = col_idx
    return headers


def _ensure_wyceny_rows(assets_path: Path, *, dry_run: bool) -> list[str]:
    messages: list[str] = []
    wyceny = pd.read_excel(assets_path, sheet_name=ASSET_EVALUATION_SHEET)
    existing = set()
    if not wyceny.empty and Properties.ID in wyceny.columns:
        existing = set(wyceny[Properties.ID].astype(str))

    if ASSET_ID in existing:
        messages.append(f"OK: {ASSET_ID!r} juz jest w {ASSET_EVALUATION_SHEET}")
        return messages

    sheets = pd.ExcelFile(assets_path).sheet_names
    if LEGACY_SHEET not in sheets:
        raise ValueError(
            f"Brak {LEGACY_SHEET!r} i brak id {ASSET_ID!r} w wycenach — nie ma skad wziac NAV"
        )

    legacy = pd.read_excel(assets_path, sheet_name=LEGACY_SHEET)
    new_rows = []
    for _, row in legacy.iterrows():
        new_rows.append(
            {
                Properties.ID: ASSET_ID,
                Properties.DATE: row["Data"],
                Properties.VALUE: row[AssetsDef.VALUE],
                Properties.CURRENCY: str(row.get("waluta", "EUR")).strip().upper() or "EUR",
                Properties.SIZE: pd.NA,
                Properties.OPERATION: OperationDomain.EVALUATION,
                Properties.UNIT_PRICE: pd.NA,
            }
        )
    added = pd.DataFrame(new_rows)
    messages.append(
        f"{'DRY-RUN' if dry_run else 'OK'}: dopisano {len(added)} wierszy do "
        f"{ASSET_EVALUATION_SHEET}: "
        + ", ".join(
            f"{pd.Timestamp(r[Properties.DATE]).date()}={float(r[Properties.VALUE])}"
            for _, r in added.iterrows()
        )
    )
    if dry_run:
        return messages

    combined = pd.concat([wyceny, added], ignore_index=True)
    with pd.ExcelWriter(
        assets_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        combined.to_excel(writer, sheet_name=ASSET_EVALUATION_SHEET, index=False)
    return messages


def _update_kind_and_drop_sheet(assets_path: Path, *, dry_run: bool) -> list[str]:
    messages: list[str] = []
    wb = load_workbook(assets_path)
    ws = wb[ASSETS_SHEET]
    headers = _header_map(ws)
    id_col = headers.get(AssetsDef.ID) or headers.get("id")
    kind_col = headers.get(AssetsDef.KIND) or headers.get("RODZAJ*")
    if id_col is None or kind_col is None:
        raise ValueError(f"Brak kolumn id/RODZAJ* w assets: {sorted(headers)}")

    kind_updated = False
    for row_idx in range(2, (ws.max_row or 1) + 1):
        asset_id = ws.cell(row_idx, id_col).value
        if asset_id is None or str(asset_id).strip() != ASSET_ID:
            continue
        old_kind = ws.cell(row_idx, kind_col).value
        messages.append(f"assets: {ASSET_ID} RODZAJ* {old_kind!r} -> {TARGET_KIND!r}")
        if not dry_run:
            ws.cell(row_idx, kind_col).value = TARGET_KIND
        kind_updated = True
        break
    if not kind_updated:
        messages.append(f"OSTRZEZENIE: brak wiersza {ASSET_ID!r} w arkuszu assets")

    if LEGACY_SHEET in wb.sheetnames:
        messages.append(f"{'DRY-RUN' if dry_run else 'OK'}: usunieto arkusz {LEGACY_SHEET!r}")
        if not dry_run:
            del wb[LEGACY_SHEET]
    else:
        messages.append(f"OK: arkusz {LEGACY_SHEET!r} juz nie istnieje")

    if not dry_run:
        wb.save(assets_path)
        messages.append(f"Zapisano {assets_path}")
    return messages


def migrate(assets_path: Path, *, dry_run: bool) -> list[str]:
    messages = _ensure_wyceny_rows(assets_path, dry_run=dry_run)
    messages.extend(_update_kind_and_drop_sheet(assets_path, dry_run=dry_run))
    return messages


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--assets", type=Path, default=None, help=f"Sciezka do {ASSETS_FILE_NAME}")
    args = parser.parse_args(argv)
    assets_path = args.assets or get_assets_file()
    for line in migrate(assets_path, dry_run=args.dry_run):
        print(line)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
