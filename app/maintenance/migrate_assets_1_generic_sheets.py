# -*- coding: utf-8 -*-
"""
Migracja assets_1.xlsx do uogolnionych nazw arkuszy/kolumn:
  zloto-monety-zakupy → inventory (moneta → instrument)
  zloto-monety-wyceny → unit-price-evaluation (moneta → instrument)
  properties-wyceny → asset-evaluation

Uzycie:
  cd app
  uv run python maintenance/migrate_assets_1_generic_sheets.py
  uv run python maintenance/migrate_assets_1_generic_sheets.py --dry-run
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from importers.assets.data_model import (
    ASSET_EVALUATION_SHEET,
    INVENTORY_SHEET,
    Inventory,
    LEGACY_ASSET_EVALUATION_SHEET,
    LEGACY_INVENTORY_SHEET,
    LEGACY_UNIT_PRICE_EVALUATION_SHEET,
    UNIT_PRICE_EVALUATION_SHEET,
)
from importers.assets.read_assets import ASSETS_FILE_NAME, get_assets_file

SHEET_RENAMES = {
    LEGACY_INVENTORY_SHEET: INVENTORY_SHEET,
    LEGACY_UNIT_PRICE_EVALUATION_SHEET: UNIT_PRICE_EVALUATION_SHEET,
    LEGACY_ASSET_EVALUATION_SHEET: ASSET_EVALUATION_SHEET,
}

INSTRUMENT_SHEETS = {INVENTORY_SHEET, UNIT_PRICE_EVALUATION_SHEET, LEGACY_INVENTORY_SHEET, LEGACY_UNIT_PRICE_EVALUATION_SHEET}
OLD_INSTRUMENT_COL = "moneta"


def migrate(assets_path: Path, *, dry_run: bool) -> list[str]:
    messages: list[str] = []
    wb = load_workbook(assets_path)

    for old, new in SHEET_RENAMES.items():
        if old in wb.sheetnames:
            if new in wb.sheetnames and new != old:
                messages.append(f"OSTRZEZENIE: {new!r} juz istnieje — pomijam rename {old!r}")
            else:
                messages.append(f"{'DRY-RUN' if dry_run else 'OK'}: arkusz {old!r} -> {new!r}")
                if not dry_run:
                    wb[old].title = new
        elif new in wb.sheetnames:
            messages.append(f"OK: arkusz {new!r} juz istnieje")
        else:
            messages.append(f"OSTRZEZENIE: brak arkusza {old!r} ani {new!r}")

    if not dry_run:
        wb.save(assets_path)

    # Rename column moneta -> instrument in inventory / unit-price sheets
    sheets = pd.ExcelFile(assets_path).sheet_names
    for sheet in sheets:
        if sheet not in (INVENTORY_SHEET, UNIT_PRICE_EVALUATION_SHEET):
            continue
        df = pd.read_excel(assets_path, sheet_name=sheet)
        if OLD_INSTRUMENT_COL not in df.columns:
            if Inventory.INSTRUMENT in df.columns:
                messages.append(f"OK: {sheet!r} ma juz kolumne {Inventory.INSTRUMENT!r}")
            else:
                messages.append(f"OSTRZEZENIE: {sheet!r} bez {OLD_INSTRUMENT_COL!r}/{Inventory.INSTRUMENT!r}")
            continue
        messages.append(
            f"{'DRY-RUN' if dry_run else 'OK'}: {sheet!r} kolumna {OLD_INSTRUMENT_COL!r} -> {Inventory.INSTRUMENT!r}"
        )
        if dry_run:
            continue
        df = df.rename(columns={OLD_INSTRUMENT_COL: Inventory.INSTRUMENT})
        with pd.ExcelWriter(
            assets_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)

    if not dry_run:
        messages.append(f"Zapisano {assets_path}")
    return messages


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--assets", type=Path, default=None, help=f"Sciezka do {ASSETS_FILE_NAME}")
    args = parser.parse_args(argv)
    assets_path = args.assets or get_assets_file()
    for line in migrate(assets_path, dry_run=args.dry_run):
        print(line)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
