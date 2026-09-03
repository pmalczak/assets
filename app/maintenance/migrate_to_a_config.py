# -*- coding: utf-8 -*-
"""
Scala assets_1.xlsx + analyse_assets_config.xlsx → a_config.xlsx (unia arkuszy),
potem usuwa stare pliki.

Użycie:
  cd app
  uv run python maintenance/migrate_to_a_config.py --dry-run
  uv run python maintenance/migrate_to_a_config.py
"""
from __future__ import annotations

import argparse
import shutil
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app_proc.data_root import (  # noqa: E402
    A_CONFIG_FILE_NAME,
    LEGACY_ANALYSE_CONFIG_FILE_NAME,
    LEGACY_ASSETS_FILE_NAME,
    get_online_data_root,
)


def _unique_table_name(target_wb, preferred: str) -> str:
    existing = set()
    for ws in target_wb.worksheets:
        if hasattr(ws, "tables"):
            existing.update(ws.tables.keys())
    name = preferred
    n = 1
    while name in existing:
        n += 1
        name = f"{preferred}_{n}"
    return name


def _copy_sheet(source_ws, target_wb, title: str) -> None:
    """Kopiuje arkusz (wartości + style komórek) do nowego workbooka."""
    target_ws = target_wb.create_sheet(title=title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
    for idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx].height = dim.height

    # Tabele Excel (best-effort): unikalne displayName w całym workbooku.
    if hasattr(source_ws, "tables") and source_ws.tables:
        from openpyxl.worksheet.table import Table

        for table in source_ws.tables.values():
            display = _unique_table_name(target_wb, table.displayName or f"Table_{title}")
            new_table = Table(displayName=display, ref=table.ref)
            if table.tableStyleInfo is not None:
                new_table.tableStyleInfo = copy(table.tableStyleInfo)
            try:
                target_ws.add_table(new_table)
            except ValueError as exc:
                print(f"  ostrzeżenie: nie skopiowano tabeli {display!r}: {exc}")


def merge_workbooks(assets_path: Path, config_path: Path, target_path: Path) -> list[str]:
    """Buduje a_config z unii arkuszy. Kolizja nazwy → ValueError."""
    wb_assets = load_workbook(assets_path)
    wb_config = load_workbook(config_path)

    names_assets = list(wb_assets.sheetnames)
    names_config = list(wb_config.sheetnames)
    collision = set(names_assets) & set(names_config)
    if collision:
        raise ValueError(
            f"Kolizja nazw arkuszy między {assets_path.name} i {config_path.name}: "
            f"{sorted(collision)}"
        )

    out = Workbook()
    # Usuń domyślny pusty arkusz
    default = out.active
    out.remove(default)

    order: list[str] = []
    for name in names_assets:
        _copy_sheet(wb_assets[name], out, name)
        order.append(name)
    for name in names_config:
        _copy_sheet(wb_config[name], out, name)
        order.append(name)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(target_path)
    return order


def smoke_read(target_path: Path) -> None:
    """Odczyt arkuszy krytycznych bez DATA_STEP / Dropbox assertów poza plikiem."""
    import pandas as pd

    from analyse_assets.config_model import (
        CATALOG_SHEET,
        MANUAL_SHEET,
        RULES_SHEET,
        AnalyseAssetsCatalog,
        AnalyseAssetsManual,
        AnalyseAssetsRules,
    )
    from importers.assets.data_model import AssetsFile

    assets = pd.read_excel(target_path, sheet_name="assets")
    AssetsFile.check_structure(assets)
    catalog = pd.read_excel(target_path, sheet_name=CATALOG_SHEET)
    AnalyseAssetsCatalog.check_structure(catalog)
    rules = pd.read_excel(target_path, sheet_name=RULES_SHEET)
    AnalyseAssetsRules.check_structure(rules)
    manual = pd.read_excel(target_path, sheet_name=MANUAL_SHEET)
    AnalyseAssetsManual.check_structure(manual)
    print(
        f"  smoke OK: assets={len(assets)} catalog={len(catalog)} "
        f"rules={len(rules)} manual={len(manual)}"
    )


def run_migration(*, dry_run: bool = False) -> None:
    root = get_online_data_root()
    assets_path = root / LEGACY_ASSETS_FILE_NAME
    config_path = root / LEGACY_ANALYSE_CONFIG_FILE_NAME
    target_path = root / A_CONFIG_FILE_NAME

    if not assets_path.is_file():
        raise FileNotFoundError(assets_path)
    if not config_path.is_file():
        raise FileNotFoundError(config_path)

    print(f"źródła: {assets_path.name} + {config_path.name}")
    print(f"cel:     {target_path}")

    if dry_run:
        wb_a = load_workbook(assets_path, read_only=True)
        wb_c = load_workbook(config_path, read_only=True)
        names_a, names_c = list(wb_a.sheetnames), list(wb_c.sheetnames)
        wb_a.close()
        wb_c.close()
        collision = set(names_a) & set(names_c)
        print(f"arkusze assets_1: {names_a}")
        print(f"arkusze config:   {names_c}")
        if collision:
            raise ValueError(f"Kolizja nazw arkuszy: {sorted(collision)}")
        print(f"unia ({len(names_a) + len(names_c)} arkuszy): {names_a + names_c}")
        print("dry-run — bez zapisu i kasowania")
        return

    if target_path.exists():
        raise FileExistsError(
            f"{target_path} już istnieje — usuń ręcznie lub przenieś, potem uruchom ponownie"
        )

    with tempfile.TemporaryDirectory() as tmp:
        staging = Path(tmp) / A_CONFIG_FILE_NAME
        order = merge_workbooks(assets_path, config_path, staging)
        print(f"zbudowano staging ({len(order)} arkuszy): {order}")
        smoke_read(staging)
        shutil.copy2(staging, target_path)
        print(f"zapisano {target_path}")

    smoke_read(target_path)

    assets_path.unlink()
    config_path.unlink()
    print(f"usunięto {LEGACY_ASSETS_FILE_NAME} i {LEGACY_ANALYSE_CONFIG_FILE_NAME}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Migracja do a_config.xlsx")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Tylko raport unii arkuszy — bez zapisu i kasowania",
    )
    args = parser.parse_args(argv)
    run_migration(dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
