# -*- coding: utf-8 -*-
"""
Dodaje kolumne source do analyse_assets_config.xlsx z zachowaniem Excel Table.

Uzycie:
  cd app
  uv run python maintenance/add_analyse_assets_source_column.py
  uv run python maintenance/add_analyse_assets_source_column.py C:/sciezka/analyse_assets_config.xlsx
"""

from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from analyse_assets.config_model import (
    CATALOG_SHEET,
    CONFIG_FILE_NAME,
    DEFAULT_TRANSACTION_SOURCE,
    RULES_SHEET,
    AnalyseAssetsCatalog,
    AnalyseAssetsRules,
)
from app_proc.data_root import get_online_data_root

_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")


def _parse_cell(cell_ref: str) -> tuple[str, int]:
    match = _CELL_REF.match(cell_ref.upper())
    if not match:
        raise ValueError(f"Niepoprawna referencja komórki: {cell_ref!r}")
    return match.group(1), int(match.group(2))


def _col_index(col_letters: str) -> int:
    value = 0
    for char in col_letters:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def _header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            headers[str(value).strip()] = col_idx
        if headers:
            return row_idx, headers
    raise ValueError(f"Brak wiersza nagłówków w arkuszu {ws.title!r}")


def _extend_tables(ws, header_row: int, new_col: int) -> None:
    for table in ws.tables.values():
        start_cell, end_cell = table.ref.split(":")
        start_col_letters, start_row = _parse_cell(start_cell)
        end_col_letters, end_row = _parse_cell(end_cell)
        start_col = _col_index(start_col_letters)
        end_col = _col_index(end_col_letters)
        if start_row != header_row:
            continue
        if new_col <= end_col:
            continue
        table.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(new_col)}{end_row}"


def _add_source_column(
    ws,
    *,
    column_name: str,
    fill_value: str | None,
) -> bool:
    header_row, headers = _header_row_and_columns(ws)
    if column_name in headers:
        return False

    new_col = max(headers.values()) + 1
    header_cell = ws.cell(header_row, new_col, value=column_name)

    # Skopiuj prosty styl z ostatniej kolumny nagłówka (jeśli jest).
    template = ws.cell(header_row, new_col - 1)
    if template.has_style:
        header_cell.font = copy(template.font)
        header_cell.fill = copy(template.fill)
        header_cell.border = copy(template.border)
        header_cell.alignment = copy(template.alignment)
        header_cell.number_format = template.number_format

    last_data_row = max(header_row, ws.max_row or header_row)
    for row_idx in range(header_row + 1, last_data_row + 1):
        # Pomiń całkowicie puste wiersze danych.
        if all(ws.cell(row_idx, col).value is None for col in headers.values()):
            continue
        cell = ws.cell(row_idx, new_col, value=fill_value if fill_value is not None else None)
        data_template = ws.cell(row_idx, new_col - 1)
        if data_template.has_style:
            cell.font = copy(data_template.font)
            cell.fill = copy(data_template.fill)
            cell.border = copy(data_template.border)
            cell.alignment = copy(data_template.alignment)

    _extend_tables(ws, header_row, new_col)
    return True


def migrate_workbook(path: Path) -> list[str]:
    wb = load_workbook(path)
    messages: list[str] = []

    if CATALOG_SHEET not in wb.sheetnames:
        raise FileNotFoundError(f"Brak arkusza {CATALOG_SHEET!r} w {path}")
    if RULES_SHEET not in wb.sheetnames:
        raise FileNotFoundError(f"Brak arkusza {RULES_SHEET!r} w {path}")

    catalog_added = _add_source_column(
        wb[CATALOG_SHEET],
        column_name=AnalyseAssetsCatalog.SOURCE,
        fill_value=DEFAULT_TRANSACTION_SOURCE,
    )
    rules_added = _add_source_column(
        wb[RULES_SHEET],
        column_name=AnalyseAssetsRules.SOURCE,
        fill_value=None,
    )

    if catalog_added:
        messages.append(f"{CATALOG_SHEET}: dodano kolumnę source={DEFAULT_TRANSACTION_SOURCE!r}")
    else:
        messages.append(f"{CATALOG_SHEET}: kolumna source już istnieje")

    if rules_added:
        messages.append(f"{RULES_SHEET}: dodano pustą kolumnę source (inherit)")
    else:
        messages.append(f"{RULES_SHEET}: kolumna source już istnieje")

    if catalog_added or rules_added:
        wb.save(path)
        messages.append(f"Zapisano: {path.resolve()}")
    else:
        messages.append("Brak zmian do zapisania.")

    return messages


def main() -> None:
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else get_online_data_root() / CONFIG_FILE_NAME
    if not target.is_file():
        raise FileNotFoundError(target)
    for line in migrate_workbook(target):
        print(line)


if __name__ == "__main__":
    main()
