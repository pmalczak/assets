# -*- coding: utf-8 -*-
"""
Migracja kolumny typ w arkuszu assets (assets_1.xlsx):
  ror          → cash_pool.ror
  cash         → investment.cash
  depozyt      → investment.depozyt
  złoto-monety → investment.złoto-monety
  udziały      → investment.udziały
  obligacje    → investment.obligacje
  property     → investment.property

Użycie:
  cd app
  uv run python maintenance/migrate_assets_typ_prefix.py
  uv run python maintenance/migrate_assets_typ_prefix.py --dry-run
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from importers.assets.data_model import AssetsFile
from importers.assets.read_assets import ASSETS_FILE_NAME, get_assets_file

ASSETS_SHEET = "assets"

TYPE_RENAMES = {
    "ror": "cash_pool.ror",
    "cash": "investment.cash",
    "depozyt": "investment.depozyt",
    "złoto-monety": "investment.złoto-monety",
    "udziały": "investment.udziały",
    "obligacje": "investment.obligacje",
    "property": "investment.property",
}


def _header_map(ws, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, (ws.max_column or 0) + 1):
        value = ws.cell(header_row, col_idx).value
        if value is None:
            continue
        headers[str(value).strip()] = col_idx
    return headers


def migrate(assets_path: Path, *, dry_run: bool) -> list[str]:
    messages: list[str] = []
    df = pd.read_excel(assets_path, sheet_name=ASSETS_SHEET)
    if AssetsFile.TYPE not in df.columns:
        raise ValueError(f"Brak kolumny {AssetsFile.TYPE!r} w arkuszu {ASSETS_SHEET!r}")

    before = df[AssetsFile.TYPE].astype(str).str.strip()
    changed = 0
    for old, new in TYPE_RENAMES.items():
        mask = before == old
        count = int(mask.sum())
        if count:
            messages.append(
                f"{'DRY-RUN' if dry_run else 'OK'}: typ {old!r} -> {new!r} ({count} wierszy)"
            )
            changed += count

    already = before.isin(TYPE_RENAMES.values())
    unknown = before[~before.isin(TYPE_RENAMES) & ~already & before.notna() & (before != "nan")]
    for value in sorted(unknown.unique()):
        messages.append(f"OSTRZEZENIE: nieznany typ {value!r} — bez mapowania")

    if changed == 0:
        messages.append("OK: brak wierszy do migracji (juz nowe wartosci lub pusto)")
        return messages

    if dry_run:
        return messages

    wb = load_workbook(assets_path)
    if ASSETS_SHEET not in wb.sheetnames:
        raise ValueError(f"Brak arkusza {ASSETS_SHEET!r} w {assets_path}")
    ws = wb[ASSETS_SHEET]
    headers = _header_map(ws)
    if AssetsFile.TYPE not in headers:
        raise ValueError(f"Brak kolumny {AssetsFile.TYPE!r} w {ASSETS_SHEET!r}")
    type_col = headers[AssetsFile.TYPE]

    for row_idx in range(2, (ws.max_row or 1) + 1):
        cell = ws.cell(row_idx, type_col)
        if cell.value is None:
            continue
        text = str(cell.value).strip()
        if text in TYPE_RENAMES:
            cell.value = TYPE_RENAMES[text]

    wb.save(assets_path)
    messages.append(f"Zapisano {assets_path}")
    return messages


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=f"Migruj kolumne typ w {ASSETS_FILE_NAME}")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--path",
        type=Path,
        default=None,
        help=f"Sciezka do {ASSETS_FILE_NAME} (domyslnie get_assets_file())",
    )
    args = parser.parse_args(argv)

    assets_path = args.path or get_assets_file()
    if not assets_path.is_file():
        raise SystemExit(f"Brak pliku: {assets_path}")

    for line in migrate(assets_path, dry_run=args.dry_run):
        print(line)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
