# -*- coding: utf-8 -*-
"""
Dashboard wartosci portfela oparty o snapshoty DATA_STEP (09 assets).

Uruchomienie:
  cd app
  uv run streamlit run app_assets.py
"""

from __future__ import annotations

import io
import re
from contextlib import redirect_stdout
from datetime import date, timedelta
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from importers.assets.data_model import AssetsDef
from importers.assets.read_assets import get_assets_file
from app_proc.calculate_assets import ASSETS_SNAPSHOT_STEP, PORTFOLIO_VALUATION_DATE
from app_proc.data_steps_root import get_data_steps_root
from app_proc.transaction_search import load_all_transactions, search_transactions
from app_proc.ui_prefs import TAB_LABELS, TABS_STATE_KEY, load_last_tab, on_tab_changed
from evaluators.valuation_date import filter_excel_rows_on_or_before
from roi.compute_roi import compute_portfolio_roi
from roi.data_model import CashFlowEvent

ROI_DISPLAY_COLUMNS = {
    "asset_id": "Aktywo",
    "capex": "Inwestycja (CAPEX)",
    "opex": "Wydatki (OPEX)",
    "revenue": "Wplywy (REVENUE)",
    "terminal_realized": "Zamkniecie (realiz.)",
    "terminal_unrealized": "Wycena (nerealiz.)",
    "roi_nominal": "ROI nominal",
    "is_sold": "Sprzedane",
}

st.set_page_config(page_title="Assets Dashboard (snapshots)", layout="wide")

SNAPSHOT_DATE_PATTERN = re.compile(r"^(\d{4}-\d{2}-\d{2})\.parquet$")
HISTORY_COLUMNS = [AssetsDef.GROUP, AssetsDef.VALUE_PLN]


def snapshots_directory() -> Path:
    return get_data_steps_root() / ASSETS_SNAPSHOT_STEP


def list_snapshot_files(snapshots_dir: Path) -> list[tuple[date, Path]]:
    if not snapshots_dir.is_dir():
        return []

    result: list[tuple[date, Path]] = []
    for path in snapshots_dir.glob("*.parquet"):
        match = SNAPSHOT_DATE_PATTERN.match(path.name)
        if not match:
            continue
        result.append((date.fromisoformat(match.group(1)), path))
    return sorted(result, key=lambda item: item[0])


def load_snapshot(path: Path) -> pd.DataFrame:
    df = pd.read_parquet(path)
    if PORTFOLIO_VALUATION_DATE not in df.columns:
        valuation_date = date.fromisoformat(path.stem)
        df = df.copy()
        df[PORTFOLIO_VALUATION_DATE] = valuation_date.isoformat()
    return df


@st.cache_data(show_spinner=False)
def load_snapshot_for_date(snapshot_date: date) -> pd.DataFrame:
    path = snapshots_directory() / f"{snapshot_date:%Y-%m-%d}.parquet"
    if not path.is_file():
        return pd.DataFrame()
    return load_snapshot(path)


@st.cache_data(show_spinner=False)
def _read_timeline_events_cached() -> pd.DataFrame:
    return _read_timeline_events()


@st.cache_data(show_spinner="Wczytywanie snapshotow...")
def build_portfolio_history_from_snapshots(
    days: int = 365,
    end_date_iso: str | None = None,
) -> dict[str, object]:
    end = date.fromisoformat(end_date_iso) if end_date_iso else date.today()
    return _build_portfolio_history_from_snapshots(days=days, end_date=end)


def _build_portfolio_history_from_snapshots(
    snapshots_dir: Path | None = None,
    days: int = 365,
    end_date: date | None = None,
) -> dict[str, object]:
    snapshots_dir = snapshots_dir or snapshots_directory()
    end = end_date or date.today()
    start = end - timedelta(days=days - 1)

    snapshot_files = list_snapshot_files(snapshots_dir)
    selected = [(snapshot_date, path) for snapshot_date, path in snapshot_files if start <= snapshot_date <= end]
    latest_target_date = selected[-1][0] if selected else None

    history_rows: list[dict[str, object]] = []
    snapshot_summaries: list[dict[str, object]] = []
    latest_snapshot = pd.DataFrame()
    latest_date: date | None = None

    for snapshot_date, path in selected:
        if snapshot_date == latest_target_date:
            assets = load_snapshot(path)
        else:
            assets = pd.read_parquet(path, columns=HISTORY_COLUMNS)
        if assets.empty:
            snapshot_summaries.append(
                {
                    "date": snapshot_date,
                    "path": str(path.name),
                    "rows": 0,
                    "total_pln": 0,
                }
            )
            continue

        by_group = (
            assets.assign(
                **{AssetsDef.VALUE_PLN: pd.to_numeric(assets[AssetsDef.VALUE_PLN], errors="coerce").fillna(0)}
            )
            .groupby(AssetsDef.GROUP, as_index=False)[AssetsDef.VALUE_PLN]
            .sum()
        )
        total_pln = float(by_group[AssetsDef.VALUE_PLN].sum())
        snapshot_summaries.append(
            {
                "date": snapshot_date,
                "path": str(path.name),
                "rows": len(assets),
                "total_pln": int(round(total_pln)),
            }
        )

        for _, row in by_group.iterrows():
            history_rows.append(
                {
                    "date": pd.Timestamp(snapshot_date),
                    "group": row[AssetsDef.GROUP],
                    "value_pln": float(row[AssetsDef.VALUE_PLN]),
                }
            )

        if latest_date is None or snapshot_date >= latest_date:
            latest_date = snapshot_date
            latest_snapshot = assets

    history = pd.DataFrame(history_rows, columns=["date", "group", "value_pln"])
    if not history.empty:
        history = history.sort_values(["date", "group"]).reset_index(drop=True)

    snapshot_by_group = pd.DataFrame(columns=["group", "value_pln"])
    snapshot_total_pln = 0.0
    if not latest_snapshot.empty:
        snapshot_by_group = (
            latest_snapshot.assign(
                **{
                    AssetsDef.VALUE_PLN: pd.to_numeric(
                        latest_snapshot[AssetsDef.VALUE_PLN],
                        errors="coerce",
                    ).fillna(0)
                }
            )
            .groupby(AssetsDef.GROUP, as_index=False)[AssetsDef.VALUE_PLN]
            .sum()
            .rename(columns={AssetsDef.GROUP: "group", AssetsDef.VALUE_PLN: "value_pln"})
        )
        snapshot_total_pln = float(snapshot_by_group["value_pln"].sum())

    return {
        "history": history,
        "snapshot_summaries": pd.DataFrame(snapshot_summaries),
        "latest_snapshot": latest_snapshot,
        "latest_snapshot_date": latest_date,
        "snapshot_by_group": snapshot_by_group,
        "snapshot_total_pln": snapshot_total_pln,
        "snapshots_dir": snapshots_dir,
        "start_date": start,
        "end_date": end,
    }


def _read_timeline_events() -> pd.DataFrame:
    try:
        assets_file = get_assets_file()
    except Exception:
        return pd.DataFrame(columns=["date", "label"])

    try:
        workbook = load_workbook(assets_file, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame(columns=["date", "label"])
    if "time-line" not in workbook.sheetnames:
        return pd.DataFrame(columns=["date", "label"])

    ws = workbook["time-line"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=["date", "label"])

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = pd.DataFrame(rows[1:], columns=header)
    if "Data" not in data.columns or "Opis" not in data.columns:
        return pd.DataFrame(columns=["date", "label"])

    data = data.rename(columns={"Data": "date", "Opis": "label"})
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data["label"] = data["label"].astype("string")
    data = data.dropna(subset=["date", "label"])
    return data[["date", "label"]].sort_values("date").reset_index(drop=True)


def render_portfolio_history(history: pd.DataFrame, timeline_events: pd.DataFrame, latest_snapshot_date: date | None):
    st.subheader("Wartosc portfela (snapshoty)")

    if history.empty:
        st.warning(
            f"Brak snapshotow w katalogu `{snapshots_directory()}`. "
            "Uruchom `maintenance/recalculate_weekly_assets_snapshots.py`."
        )
        return

    available_groups = sorted(history["group"].dropna().unique().tolist())
    selected_groups = st.multiselect(
        "Widoczne grupy",
        options=available_groups,
        default=available_groups,
    )
    if not selected_groups:
        st.warning("Wybierz przynajmniej jedna grupe.")
        return

    visible_history = history[history["group"].isin(selected_groups)].copy()
    totals = visible_history.groupby("date", as_index=False)["value_pln"].sum().sort_values("date")
    current_value = float(totals["value_pln"].iloc[-1])
    start_value = float(totals["value_pln"].iloc[0])
    delta_value = current_value - start_value
    delta_pct = (delta_value / start_value * 100) if start_value else 0.0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Ostatni snapshot", f"{current_value:,.0f} PLN".replace(",", " "))
    c2.metric("Pierwszy w oknie", f"{start_value:,.0f} PLN".replace(",", " "))
    c3.metric("Zmiana w oknie", f"{delta_value:,.0f} PLN".replace(",", " "), f"{delta_pct:+.1f}%")
    if latest_snapshot_date:
        c4.metric("Data ostatniego snapshotu", latest_snapshot_date.isoformat())
    else:
        c4.metric("Data ostatniego snapshotu", "-")

    chart_data = (
        visible_history.groupby(["date", "group"], as_index=False)["value_pln"]
        .sum()
        .sort_values(["date", "group"])
    )
    month_lines = pd.DataFrame(
        {
            "date": pd.date_range(
                chart_data["date"].min().normalize(),
                chart_data["date"].max().normalize(),
                freq="MS",
            )
        }
    )

    area_chart = (
        alt.Chart(chart_data)
        .mark_area(interpolate="step-after")
        .encode(
            x=alt.X("date:T", title="Data"),
            y=alt.Y("value_pln:Q", title="Wartosc portfela (PLN)", stack="zero"),
            color=alt.Color("group:N", title="Grupa"),
            order=alt.Order("group:N"),
            tooltip=[
                alt.Tooltip("date:T", title="Data"),
                alt.Tooltip("group:N", title="Grupa"),
                alt.Tooltip("value_pln:Q", title="Wartosc", format=",.0f"),
            ],
        )
    )
    month_rules = (
        alt.Chart(month_lines)
        .mark_rule(color="#666666", strokeWidth=1, opacity=0.35)
        .encode(x="date:T")
    )
    chart = area_chart + month_rules

    if not timeline_events.empty:
        visible_events = timeline_events[
            (timeline_events["date"] >= chart_data["date"].min())
            & (timeline_events["date"] <= chart_data["date"].max())
        ].copy()

        if not visible_events.empty:
            totals_with_events = totals.rename(columns={"value_pln": "total_value_pln"})
            visible_events = visible_events.merge(totals_with_events, on="date", how="left")
            visible_events["total_value_pln"] = visible_events["total_value_pln"].ffill()
            visible_events["label_y"] = visible_events["total_value_pln"] * 1.01

            event_rules = (
                alt.Chart(visible_events)
                .mark_rule(color="#f4d03f", strokeWidth=2, opacity=0.7)
                .encode(
                    x="date:T",
                    tooltip=[alt.Tooltip("date:T", title="Data"), alt.Tooltip("label:N", title="Opis")],
                )
            )
            event_points = (
                alt.Chart(visible_events)
                .mark_point(color="#f4d03f", filled=True, size=90, shape="diamond")
                .encode(
                    x="date:T",
                    y=alt.Y("total_value_pln:Q"),
                    tooltip=[alt.Tooltip("date:T", title="Data"), alt.Tooltip("label:N", title="Opis")],
                )
            )
            event_labels = (
                alt.Chart(visible_events)
                .mark_text(color="#f4d03f", angle=325, align="left", baseline="bottom", dx=6, dy=-4, fontSize=11)
                .encode(
                    x="date:T",
                    y=alt.Y("label_y:Q"),
                    text="label:N",
                )
            )
            chart = chart + event_rules + event_points + event_labels

    st.altair_chart(chart, use_container_width=True)
    st.caption(
        f"Wykres budowany wylacznie ze snapshotow `{ASSETS_SNAPSHOT_STEP}` "
        f"(kolumna `{PORTFOLIO_VALUATION_DATE}`). Kazdy punkt to wartosc portfela "
        "wyliczona przez `calculate_assets()` na dana date wyceny — bez rekonstrukcji z transakcji."
    )


@st.cache_data(show_spinner="Wczytywanie transakcji...")
def _load_transactions_cached() -> pd.DataFrame:
    return load_all_transactions()


def render_transaction_search() -> None:
    st.subheader("Wyszukiwanie transakcji")

    try:
        transactions = _load_transactions_cached()
    except Exception as exc:
        st.error("Nie udalo sie wczytac transakcji.")
        st.exception(exc)
        return

    st.caption(
        f"Przeszukiwane zrodla: mBank i Revolut ({len(transactions):,} transakcji). "
        "Szukamy w polach: opis, tytul, kontrahent, konto."
    )

    query = st.text_input(
        "Szukaj",
        placeholder="np. nazwa kontrahenta, fragment tytulu, opis operacji...",
        key="transaction_search_query",
    )
    case_sensitive = st.checkbox("Uwzgledniaj wielkosc liter", value=False, key="transaction_search_case")

    col1, col2, col3 = st.columns(3)
    col1.metric("Wszystkie transakcje", f"{len(transactions):,}".replace(",", " "))
    col2.metric("Konta / zrodla", transactions["asset_id"].nunique() if not transactions.empty else 0)

    if not query.strip():
        st.info("Wpisz fraze wyszukiwania, aby zobaczyc pasujace transakcje.")
        return

    results = search_transactions(transactions, query, case_sensitive=case_sensitive)
    col3.metric("Wyniki", len(results))

    if results.empty:
        st.warning(f"Brak transakcji zawierajacych „{query}” w polach tekstowych.")
        return

    st.dataframe(results, use_container_width=True, hide_index=True, height=520)

    csv = results.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="Pobierz wyniki (CSV)",
        data=csv,
        file_name="transakcje_wyszukiwanie.csv",
        mime="text/csv",
        key="transaction_search_csv",
    )


@st.cache_data(show_spinner="Liczenie ROI...")
def _load_roi_data(valuation_date_iso: str) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    valuation_date = date.fromisoformat(valuation_date_iso)
    summary, events_by_asset = compute_portfolio_roi(valuation_date, use_cache=True)
    return summary, events_by_asset


def render_roi(default_valuation_date: date | None) -> None:
    st.subheader("ROI nieruchomosci")

    valuation_date = st.date_input(
        "Data wyceny ROI",
        value=default_valuation_date or date.today(),
        key="roi_valuation_date",
    )

    try:
        summary, events_by_asset = _load_roi_data(valuation_date.isoformat())
    except Exception as exc:
        st.error("Nie udalo sie policzyc ROI.")
        st.exception(exc)
        return

    if summary.empty:
        st.info("Brak danych ROI w katalogu analyse_assets.")
        return

    st.caption(
        "ROI nominalny = suma alokowanych przeplywow + wycena z arkusza properties dla otwartych inwestycji. "
        "Konfiguracja: `analyse_assets/analyse_assets_config.xlsx`."
    )

    total_roi = int(summary["roi_nominal"].sum())
    sold_count = int(summary["is_sold"].sum())
    c1, c2, c3 = st.columns(3)
    c1.metric("Liczba aktywow", len(summary))
    c2.metric("Sprzedane", sold_count)
    c3.metric("Suma ROI nominal", f"{total_roi:,}".replace(",", " "))

    display = summary[list(ROI_DISPLAY_COLUMNS.keys())].rename(columns=ROI_DISPLAY_COLUMNS)
    for col in ROI_DISPLAY_COLUMNS.values():
        if col == "Sprzedane":
            continue
        if col in display.columns:
            display[col] = display[col].map(
                lambda v: f"{v:,}".replace(",", " ") if isinstance(v, (int, float)) else v
            )

    st.dataframe(display, use_container_width=True, hide_index=True)

    csv = summary.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="Pobierz ROI (CSV)",
        data=csv,
        file_name=f"roi_{valuation_date:%Y-%m-%d}.csv",
        mime="text/csv",
        key="roi_csv_download",
    )

    warned = summary[summary["warnings"].astype(str).str.len() > 0]
    if not warned.empty:
        st.markdown("**Ostrzezenia**")
        st.dataframe(warned[["asset_id", "warnings"]], use_container_width=True, hide_index=True)

    st.markdown("**Szczegoly przeplywow**")
    asset_ids = sorted(summary["asset_id"].astype(str).tolist())
    selected_asset = st.selectbox("Aktywo", options=asset_ids, key="roi_selected_asset")
    events = events_by_asset.get(selected_asset, pd.DataFrame(columns=list(CashFlowEvent.COLUMN_ORDER)))
    if events.empty:
        st.info("Brak zarejestrowanych przeplywow dla tego aktywa.")
    else:
        events_display = filter_excel_rows_on_or_before(events, CashFlowEvent.DATE, valuation_date)
        st.dataframe(events_display, use_container_width=True, hide_index=True, height=280)


def render_main_reports(snapshot_date: date | None, assets: pd.DataFrame):
    from asset_reports import rap1, rap2

    st.subheader("Raporty (jak w main.py)")

    snapshot_files = list_snapshot_files(snapshots_directory())
    if not snapshot_files:
        st.warning(
            f"Brak snapshotow w katalogu `{snapshots_directory()}`. "
            "Uruchom `maintenance/recalculate_weekly_assets_snapshots.py`."
        )
        return

    available_dates = [item[0] for item in snapshot_files]
    default_index = len(available_dates) - 1
    if snapshot_date in available_dates:
        default_index = available_dates.index(snapshot_date)

    selected_date = st.selectbox(
        "Data snapshotu",
        options=available_dates,
        index=default_index,
        format_func=lambda d: d.isoformat(),
    )
    if selected_date != snapshot_date:
        assets = load_snapshot_for_date(selected_date)

    if assets.empty:
        st.warning(f"Brak danych w snapshotcie {selected_date:%Y-%m-%d}.")
        return

    st.caption(f"Zrodlo: `{ASSETS_SNAPSHOT_STEP}/{selected_date:%Y-%m-%d}.parquet`")

    st.markdown("**Pelna lista aktywow**")
    st.dataframe(assets, use_container_width=True, hide_index=True, height=360)

    st.markdown("**RAP 2**")
    rap2_buffer = io.StringIO()
    with redirect_stdout(rap2_buffer):
        rap2(assets)
    st.code(rap2_buffer.getvalue().strip(), language=None)

    st.markdown("**RAP 1**")
    st.code(rap1(assets).to_string(col_space=15), language=None)


def render_diagnostics(data: dict[str, object]):
    st.divider()
    st.subheader("Diagnostyka snapshotow")

    summaries = data["snapshot_summaries"]
    history = data["history"]

    c1, c2, c3 = st.columns(3)
    c1.metric("Snapshotow w oknie", len(summaries))
    c2.metric("Punktow wykresu", len(history))
    c3.metric("Grup", history["group"].nunique() if not history.empty else 0)

    st.markdown(f"**Katalog:** `{data['snapshots_dir']}`")
    st.markdown("**Lista snapshotow**")
    if summaries.empty:
        st.info("Brak plikow parquet w katalogu snapshotow.")
    else:
        display = summaries.copy()
        display["date"] = display["date"].apply(lambda d: d.isoformat() if isinstance(d, date) else d)
        st.dataframe(display.sort_values("date"), use_container_width=True, hide_index=True)

    st.markdown("**Ostatni snapshot per grupa**")
    st.dataframe(
        data["snapshot_by_group"].sort_values("group"),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("**Historia per grupa (z snapshotow)**")
    st.dataframe(
        history.sort_values(["date", "group"]),
        use_container_width=True,
        hide_index=True,
        height=280,
    )


def build_data(days: int = 365) -> dict[str, object]:
    data = build_portfolio_history_from_snapshots(days=days, end_date_iso=date.today().isoformat())
    data["timeline_events"] = _read_timeline_events_cached()

    latest_snapshot = data["latest_snapshot"]
    excel_buffer = io.BytesIO()
    if isinstance(latest_snapshot, pd.DataFrame) and not latest_snapshot.empty:
        latest_snapshot.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)
    data["excel_bytes"] = excel_buffer
    return data


def main():
    st.title("Assets Dashboard (snapshoty DATA_STEP)")

    with st.spinner("Ladowanie snapshotow..."):
        try:
            data = build_data()
        except Exception as e:
            st.error("Wystapil blad podczas wczytywania snapshotow.")
            st.exception(e)
            return

    left, right = st.columns([1, 1])
    with left:
        label = "Ostatni snapshot"
        if data["latest_snapshot_date"]:
            label += f" ({data['latest_snapshot_date']:%Y-%m-%d})"
        st.metric(label, f"{data['snapshot_total_pln']:,.0f} PLN".replace(",", " "))
    with right:
        file_name = "assets_evaluation.xlsx"
        if data["latest_snapshot_date"]:
            file_name = f"assets_evaluation_{data['latest_snapshot_date']:%Y-%m-%d}.xlsx"
        st.download_button(
            label="Pobierz ostatni snapshot (xlsx)",
            data=data["excel_bytes"],
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=data["latest_snapshot"].empty,
        )

    if TABS_STATE_KEY not in st.session_state:
        st.session_state[TABS_STATE_KEY] = load_last_tab()

    tab_chart, tab_reports, tab_search, tab_roi = st.tabs(
        TAB_LABELS,
        key=TABS_STATE_KEY,
        default=st.session_state[TABS_STATE_KEY],
        on_change=on_tab_changed,
    )

    with tab_chart:
        render_portfolio_history(
            data["history"],
            data["timeline_events"],
            data["latest_snapshot_date"],
        )
        render_diagnostics(data)

    with tab_reports:
        render_main_reports(data["latest_snapshot_date"], data["latest_snapshot"])

    with tab_search:
        render_transaction_search()

    with tab_roi:
        render_roi(data["latest_snapshot_date"])


if __name__ == "__main__":
    pd.options.future.infer_string = True
    main()
