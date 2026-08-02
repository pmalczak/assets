# -*- coding: utf-8 -*-
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from analyse_assets.config_model import CATALOG_SHEET, MANUAL_SHEET, RULES_SHEET
from app_proc.data_root import A_CONFIG_FILE_NAME
from maintenance.migrate_to_a_config import merge_workbooks, smoke_read


def _write_assets_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "assets"
    ws.append(["id", "typ", "grupa", "opis", "RODZAJ*", "waluta", "dostęp"])
    ws.append(
        [
            "cash",
            "investment.cash",
            "0 gotówka",
            "test",
            "assets.cash",
            "EUR",
            "",
        ]
    )
    inv = wb.create_sheet("inventory")
    inv.append(["Data", "instrument", "waga", "sztuki", "notatki"])
    wb.save(path)


def _write_config_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = CATALOG_SHEET
    ws.append(["asset_id", "output_file", "order", "enabled", "pool_id"])
    ws.append(["cash", "a_cash.xlsx", 1, 1, "mbank_eur"])
    rules = wb.create_sheet(RULES_SHEET)
    rules.append(
        [
            "asset_id",
            "step_id",
            "step_order",
            "mapping",
            "condition_group",
            "pool_id",
            "field",
            "operator",
            "value",
            "Uwagi",
        ]
    )
    manual = wb.create_sheet(MANUAL_SHEET)
    manual.append(
        ["asset_id", "step_order", "date", "amount", "category", "description"]
    )
    wb.save(path)


class MergeToAConfigTests(unittest.TestCase):
    def test_merge_union_and_collision(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            assets = root / "assets_1.xlsx"
            config = root / "analyse_assets_config.xlsx"
            target = root / A_CONFIG_FILE_NAME
            _write_assets_xlsx(assets)
            _write_config_xlsx(config)

            order = merge_workbooks(assets, config, target)
            self.assertIn("assets", order)
            self.assertIn(CATALOG_SHEET, order)
            self.assertTrue(target.is_file())
            smoke_read(target)

            # collision
            bad = root / "bad.xlsx"
            _write_assets_xlsx(bad)
            with self.assertRaises(ValueError):
                merge_workbooks(assets, bad, root / "out.xlsx")


if __name__ == "__main__":
    unittest.main()
