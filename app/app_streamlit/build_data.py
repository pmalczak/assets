from __future__ import annotations

import io
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from app_proc.recalculate_snapshots import PORTFOLIO_WINDOW_DAYS
from app_proc.snapshots import snapshots_directory, list_snapshot_files, load_snapshot

from importers.assets.data_model import AssetsDef
from importers.assets.read_assets import get_assets_file

HISTORY_COLUMNS = [AssetsDef.GROUP, AssetsDef.VALUE_PLN]


def build_data(days: int = PORTFOLIO_WINDOW_DAYS) -> dict[str, object]:
    data = build_portfolio_history_from_snapshots(
        days=days,
        end_date_iso=date.today().isoformat(),
        _schema=2,
    )
    data["timeline_events"] = _read_timeline_events_cached()

    latest_snapshot = data["latest_snapshot"]
    excel_buffer = io.BytesIO()
    if isinstance(latest_snapshot, pd.DataFrame) and not latest_snapshot.empty:
        latest_snapshot.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)
    data["excel_bytes"] = excel_buffer
    return data


@st.cache_data(show_spinner=False)
def _read_timeline_events_cached() -> pd.DataFrame:
    return _read_timeline_events()


@st.cache_data(show_spinner="Wczytywanie snapshotow...")
def build_portfolio_history_from_snapshots(
    days: int = PORTFOLIO_WINDOW_DAYS,
    end_date_iso: str | None = None,
    _schema: int = 2,
) -> dict[str, object]:
    del _schema  # tylko do uniewazniania cache przy zmianie ksztaltu wyniku
    end = date.fromisoformat(end_date_iso) if end_date_iso else date.today()
    return _build_portfolio_history_from_snapshots(days=days, end_date=end)


def _build_portfolio_history_from_snapshots(
    snapshots_dir: Path | None = None,
    days: int = PORTFOLIO_WINDOW_DAYS,
    end_date: date | None = None,
) -> dict[str, object]:
    snapshots_dir = snapshots_dir or snapshots_directory()
    end = end_date or date.today()
    start = end - timedelta(days=days - 1)

    snapshot_files = list_snapshot_files(snapshots_dir)
    selected = [(snapshot_date, path) for snapshot_date, path in snapshot_files if start <= snapshot_date <= end]
    latest_target_date = selected[-1][0] if selected else None

    history_rows: list[dict[str, object]] = []
    snapshot_summaries: list[dict[str, object]] = []
    latest_snapshot = pd.DataFrame()
    latest_date: date | None = None

    for snapshot_date, path in selected:
        if snapshot_date == latest_target_date:
            assets = load_snapshot(path)
        else:
            assets = pd.read_parquet(path, columns=HISTORY_COLUMNS)
        if assets.empty:
            snapshot_summaries.append(
                {
                    "date": snapshot_date,
                    "path": str(path.name),
                    "rows": 0,
                    "total_pln": 0,
                }
            )
            continue

        by_group = (
            assets.assign(
                **{AssetsDef.VALUE_PLN: pd.to_numeric(assets[AssetsDef.VALUE_PLN], errors="coerce").fillna(0)}
            )
            .groupby(AssetsDef.GROUP, as_index=False)[AssetsDef.VALUE_PLN]
            .sum()
        )
        total_pln = float(by_group[AssetsDef.VALUE_PLN].sum())
        snapshot_summaries.append(
            {
                "date": snapshot_date,
                "path": str(path.name),
                "rows": len(assets),
                "total_pln": int(round(total_pln)),
            }
        )

        for _, row in by_group.iterrows():
            history_rows.append(
                {
                    "date": pd.Timestamp(snapshot_date),
                    "group": row[AssetsDef.GROUP],
                    "value_pln": float(row[AssetsDef.VALUE_PLN]),
                }
            )

        if latest_date is None or snapshot_date >= latest_date:
            latest_date = snapshot_date
            latest_snapshot = assets

    history = pd.DataFrame(history_rows, columns=["date", "group", "value_pln"])
    if not history.empty:
        history = history.sort_values(["date", "group"]).reset_index(drop=True)

    snapshot_by_type = pd.DataFrame(columns=["type", "value_pln"])
    snapshot_total_pln = 0.0
    if not latest_snapshot.empty:
        snapshot_by_type = (
            latest_snapshot.assign(
                **{
                    AssetsDef.VALUE_PLN: pd.to_numeric(
                        latest_snapshot[AssetsDef.VALUE_PLN],
                        errors="coerce",
                    ).fillna(0)
                }
            )
            .groupby(AssetsDef.TYPE, as_index=False)[AssetsDef.VALUE_PLN]
            .sum()
            .rename(columns={AssetsDef.TYPE: "type", AssetsDef.VALUE_PLN: "value_pln"})
        )
        snapshot_total_pln = float(snapshot_by_type["value_pln"].sum())

    return {
        "history": history,
        "snapshot_summaries": pd.DataFrame(snapshot_summaries),
        "latest_snapshot": latest_snapshot,
        "latest_snapshot_date": latest_date,
        "snapshot_by_type": snapshot_by_type,
        "snapshot_total_pln": snapshot_total_pln,
        "snapshots_dir": snapshots_dir,
        "start_date": start,
        "end_date": end,
    }


def _read_timeline_events() -> pd.DataFrame:
    try:
        assets_file = get_assets_file()
    except Exception:
        return pd.DataFrame(columns=["date", "label"])

    try:
        workbook = load_workbook(assets_file, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame(columns=["date", "label"])
    if "time-line" not in workbook.sheetnames:
        return pd.DataFrame(columns=["date", "label"])

    ws = workbook["time-line"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=["date", "label"])

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = pd.DataFrame(rows[1:], columns=header)
    if "Data" not in data.columns or "Opis" not in data.columns:
        return pd.DataFrame(columns=["date", "label"])

    data = data.rename(columns={"Data": "date", "Opis": "label"})
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data["label"] = data["label"].astype("string")
    data = data.dropna(subset=["date", "label"])
    return data[["date", "label"]].sort_values("date").reset_index(drop=True)
