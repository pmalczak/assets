import io
import sys
from pathlib import Path

import pandas as pd
import streamlit as st
import altair as alt
from openpyxl import load_workbook

project_root = Path().home() / r"PycharmProjects\github_common_py"
sys.path.append(str(project_root))

from main_proc.check_wrong_catalogs import check_wrong_catalogs
from main_proc.data_root import get_online_data_root
from data_step.data_step import DATA_STEP
from evaluators.evaluate_assets import evaluate_assets
from fx.data_model import LastFx
from importers.assets.data_model import AssetsDef, AssetsFile
from importers.assets.read_assets import get_assets_file, read_assets
from nbp_fx_repo.nbp_fx_repository import NBP_API_EUR, NbpFxRepository
from main_app_proc.portfolio_history import build_portfolio_history

st.set_page_config(page_title="Assets Dashboard", layout="wide")


def build_data():
    local_data_steps_root = Path(__file__).parent.parent
    DATA_STEP.init_steps(root=local_data_steps_root)

    data_root = get_online_data_root()

    metadata_root: Path = DATA_STEP.metadata.get_metadata_root() / "fx"
    metadata_root.mkdir(parents=True, exist_ok=True)

    fx_repo = NbpFxRepository(target_directory=metadata_root, min_year=2005)
    fx_rates = fx_repo.update_to_date()
    fx_rates = fx_rates[[NBP_API_EUR]]

    assets_catalog = read_assets()
    check_wrong_catalogs(data_root, assets_catalog)
    history_data = build_portfolio_history(assets_catalog, fx_rates)

    assets = evaluate_assets(data_root, assets_catalog, fx_rates)
    assets = assets.sort_values(by=[AssetsFile.GROUP, AssetsFile.ID])
    assets = assets[assets[AssetsDef.VALUE] != 0]
    assets = assets.drop(columns=[AssetsDef.NOTES, LastFx.FX])
    snapshot_by_group = (
        assets[[AssetsDef.GROUP, AssetsDef.VALUE_PLN]]
        .assign(**{AssetsDef.VALUE_PLN: pd.to_numeric(assets[AssetsDef.VALUE_PLN], errors="coerce").fillna(0)})
        .groupby(AssetsDef.GROUP, as_index=False)[AssetsDef.VALUE_PLN]
        .sum()
        .rename(columns={AssetsDef.GROUP: "group", AssetsDef.VALUE_PLN: "value_pln"})
    )
    snapshot_total_pln = float(snapshot_by_group["value_pln"].sum())
    history, offsets = _align_history_to_snapshot(history_data["history"], snapshot_by_group)

    excel_buffer = io.BytesIO()
    assets.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    return {
        "excel_bytes": excel_buffer,
        "portfolio_history_raw": history_data["history"],
        "portfolio_history": history,
        "portfolio_offsets": offsets,
        "snapshot_by_group": snapshot_by_group,
        "timeline_events": _read_timeline_events(),
        "history_skipped_assets": history_data["skipped_assets"],
        "snapshot_total_pln": snapshot_total_pln,
    }


def _read_timeline_events() -> pd.DataFrame:
    workbook = load_workbook(get_assets_file(), read_only=True, data_only=True)
    if "time-line" not in workbook.sheetnames:
        return pd.DataFrame(columns=["date", "label"])

    ws = workbook["time-line"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=["date", "label"])

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = pd.DataFrame(rows[1:], columns=header)
    if "Data" not in data.columns or "Opis" not in data.columns:
        return pd.DataFrame(columns=["date", "label"])

    data = data.rename(columns={"Data": "date", "Opis": "label"})
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data["label"] = data["label"].astype("string")
    data = data.dropna(subset=["date", "label"])
    return data[["date", "label"]].sort_values("date").reset_index(drop=True)


def _align_history_to_snapshot(history: pd.DataFrame, snapshot_by_group: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    aligned = history.copy()
    today = pd.Timestamp.today().normalize()
    snapshot_map = snapshot_by_group.set_index("group")["value_pln"]

    if aligned.empty:
        date_range = pd.date_range(end=today, periods=365, freq="D")
        fallback_rows = [
            {
                "date": date,
                "group": row["group"],
                "value_pln": row["value_pln"] if date == today else 0.0,
            }
            for date in date_range
            for _, row in snapshot_by_group.iterrows()
        ]
        fallback = pd.DataFrame(fallback_rows)
        offsets = snapshot_by_group.copy()
        offsets["history_value_pln"] = 0.0
        offsets["missing_constant_value"] = offsets["value_pln"]
        return fallback[["date", "group", "value_pln"]], offsets[["group", "value_pln", "history_value_pln", "missing_constant_value"]]

    aligned["date"] = pd.to_datetime(aligned["date"])
    aligned = aligned.sort_values("date").reset_index(drop=True)
    aligned["value_pln"] = pd.to_numeric(aligned["value_pln"], errors="coerce").fillna(0)

    first_date = aligned["date"].min()
    last_date = aligned["date"].max()
    if last_date < today:
        full_dates = pd.date_range(first_date, today, freq="D")
    else:
        full_dates = pd.date_range(first_date, last_date, freq="D")

    last_history_by_group = (
        aligned[aligned["date"] == last_date][["group", "value_pln"]]
        .groupby("group", as_index=False)["value_pln"]
        .sum()
        .rename(columns={"value_pln": "history_value_pln"})
    )

    offsets = snapshot_by_group.merge(last_history_by_group, on="group", how="outer").fillna(0)
    offsets["missing_constant_value"] = offsets["value_pln"] - offsets["history_value_pln"]

    if last_date < today:
        last_raw = aligned[aligned["date"] == last_date][["group", "value_pln"]].copy()
        extension_rows = []
        for date in pd.date_range(last_date + pd.Timedelta(days=1), today, freq="D"):
            x = last_raw.copy()
            x["date"] = date
            extension_rows.append(x)
        if extension_rows:
            aligned = pd.concat([aligned] + extension_rows, ignore_index=True)

    missing_groups = sorted(set(snapshot_by_group["group"]) - set(aligned["group"].unique()))
    if missing_groups:
        missing_rows = [
            {
                "date": date,
                "group": group,
                "value_pln": float(snapshot_map[group]) if date == today else 0.0,
            }
            for date in full_dates
            for group in missing_groups
        ]
        aligned = pd.concat([aligned, pd.DataFrame(missing_rows)], ignore_index=True)

    today_rows = []
    for group, snap_value in snapshot_map.items():
        group_today_mask = (aligned["date"] == today) & (aligned["group"] == group)
        if group_today_mask.any():
            aligned.loc[group_today_mask, "value_pln"] = snap_value
        else:
            today_rows.append({"date": today, "group": group, "value_pln": float(snap_value)})
    if today_rows:
        aligned = pd.concat([aligned, pd.DataFrame(today_rows)], ignore_index=True)

    return (
        aligned.sort_values(["date", "group"]).reset_index(drop=True),
        offsets[["group", "value_pln", "history_value_pln", "missing_constant_value"]].sort_values("group").reset_index(drop=True),
    )


def render_portfolio_history(history: pd.DataFrame, skipped_assets: list[str], timeline_events: pd.DataFrame):
    st.subheader("Wartosc portfela za ostatni rok")

    if history.empty:
        st.warning("Brak danych historycznych do zbudowania wykresu portfela.")
    else:
        available_groups = sorted(history["group"].dropna().unique().tolist())
        selected_groups = st.multiselect(
            "Widoczne grupy",
            options=available_groups,
            default=available_groups,
        )
        if not selected_groups:
            st.warning("Wybierz przynajmniej jedna grupe.")
            return

        visible_history = history[history["group"].isin(selected_groups)].copy()
        totals = visible_history.groupby("date", as_index=False)["value_pln"].sum().sort_values("date")
        current_value = float(totals["value_pln"].iloc[-1])
        start_value = float(totals["value_pln"].iloc[0])
        delta_value = current_value - start_value
        delta_pct = (delta_value / start_value * 100) if start_value else 0.0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Biezaca wartosc", f"{current_value:,.0f} PLN".replace(",", " "))
        c2.metric("Wartosc 12M temu", f"{start_value:,.0f} PLN".replace(",", " "))
        c3.metric("Zmiana 12M", f"{delta_value:,.0f} PLN".replace(",", " "), f"{delta_pct:+.1f}%")
        c4.metric("Zakres dat", totals["date"].iloc[0].strftime("%Y-%m-%d"))

        chart_data = (
            visible_history.groupby(["date", "group"], as_index=False)["value_pln"]
            .sum()
            .sort_values(["date", "group"])
        )
        month_lines = pd.DataFrame(
            {
                "date": pd.date_range(
                    chart_data["date"].min().normalize(),
                    chart_data["date"].max().normalize(),
                    freq="MS",
                )
            }
        )

        area_chart = (
            alt.Chart(chart_data)
            .mark_area()
            .encode(
                x=alt.X("date:T", title="Data"),
                y=alt.Y("value_pln:Q", title="Wartosc portfela (PLN)", stack="zero"),
                color=alt.Color("group:N", title="Grupa"),
                order=alt.Order("group:N"),
                tooltip=[
                    alt.Tooltip("date:T", title="Data"),
                    alt.Tooltip("group:N", title="Grupa"),
                    alt.Tooltip("value_pln:Q", title="Wartosc", format=",.0f"),
                ],
            )
        )
        month_rules = (
            alt.Chart(month_lines)
            .mark_rule(color="#666666", strokeWidth=1, opacity=0.35)
            .encode(x="date:T")
        )
        chart = area_chart + month_rules

        if not timeline_events.empty:
            visible_events = timeline_events[
                (timeline_events["date"] >= chart_data["date"].min()) &
                (timeline_events["date"] <= chart_data["date"].max())
            ].copy()

            if not visible_events.empty:
                totals_with_events = totals.rename(columns={"value_pln": "total_value_pln"})
                visible_events = visible_events.merge(totals_with_events, on="date", how="left")
                visible_events["label_y"] = visible_events["total_value_pln"] * 1.01

                event_rules = (
                    alt.Chart(visible_events)
                    .mark_rule(color="#f4d03f", strokeWidth=2, opacity=0.7)
                    .encode(x="date:T", tooltip=[alt.Tooltip("date:T", title="Data"), alt.Tooltip("label:N", title="Opis")])
                )
                event_points = (
                    alt.Chart(visible_events)
                    .mark_point(color="#f4d03f", filled=True, size=90, shape="diamond")
                    .encode(
                        x="date:T",
                        y=alt.Y("total_value_pln:Q"),
                        tooltip=[alt.Tooltip("date:T", title="Data"), alt.Tooltip("label:N", title="Opis")],
                    )
                )
                event_labels = (
                    alt.Chart(visible_events)
                    .mark_text(color="#f4d03f", angle=325, align="left", baseline="bottom", dx=6, dy=-4, fontSize=11)
                    .encode(
                        x="date:T",
                        y=alt.Y("label_y:Q"),
                        text="label:N",
                    )
                )
                chart = chart + event_rules + event_points + event_labels

        st.altair_chart(chart, use_container_width=True)

        st.caption(
            "Historia jest odtwarzana z danych zrodlowych i kursow NBP dla EUR. "
            "Biezacy snapshot jest uzgadniany tylko na ostatni dzien wykresu; wczesniejsze punkty pokazuja "
            "wartosc wyliczona z historii transakcji (bez retroaktywnego doliczania roznicy do snapshotu)."
        )

    if skipped_assets:
        st.info(
            "Pominiete lub niepelne zrodla historii: "
            + ", ".join(skipped_assets[:8])
            + (" ..." if len(skipped_assets) > 8 else "")
        )


def render_diagnostics(
    raw_history: pd.DataFrame,
    aligned_history: pd.DataFrame,
    offsets: pd.DataFrame,
    snapshot_by_group: pd.DataFrame,
):
    st.divider()
    st.subheader("Diagnostyka")

    raw_totals = (
        raw_history.groupby("date", as_index=False)["value_pln"].sum().sort_values("date")
        if not raw_history.empty
        else pd.DataFrame(columns=["date", "value_pln"])
    )
    aligned_totals = (
        aligned_history.groupby("date", as_index=False)["value_pln"].sum().sort_values("date")
        if not aligned_history.empty
        else pd.DataFrame(columns=["date", "value_pln"])
    )

    c1, c2, c3 = st.columns(3)
    c1.metric("Punktow raw", len(raw_history))
    c2.metric("Punktow aligned", len(aligned_history))
    c3.metric("Grup", aligned_history["group"].nunique() if not aligned_history.empty else 0)

    st.markdown("**Offsety dodane do grup**")
    st.dataframe(offsets, use_container_width=True, hide_index=True)

    st.markdown("**Snapshot per grupa**")
    st.dataframe(snapshot_by_group.sort_values("group"), use_container_width=True, hide_index=True)

    st.markdown("**Suma portfela po dniach: raw vs aligned**")
    totals_compare = raw_totals.rename(columns={"value_pln": "raw_total_pln"}).merge(
        aligned_totals.rename(columns={"value_pln": "aligned_total_pln"}),
        on="date",
        how="outer",
    ).sort_values("date")
    st.dataframe(totals_compare, use_container_width=True, hide_index=True)

    st.markdown("**Surowa historia per grupa**")
    st.dataframe(
        raw_history.sort_values(["date", "group"]),
        use_container_width=True,
        hide_index=True,
        height=280,
    )

    st.markdown("**Historia po uzgodnieniu per grupa**")
    st.dataframe(
        aligned_history.sort_values(["date", "group"]),
        use_container_width=True,
        hide_index=True,
        height=280,
    )


def main():
    st.title("Assets Dashboard")

    with st.spinner("Ladowanie i przetwarzanie danych..."):
        try:
            data = build_data()
        except Exception as e:
            st.error("Wystapil blad podczas przygotowywania danych.")
            st.exception(e)
            return

    left, right = st.columns([1, 1])
    with left:
        st.metric("Biezacy snapshot", f"{data['snapshot_total_pln']:,.0f} PLN".replace(",", " "))
    with right:
        st.download_button(
            label="Pobierz assets_evaluation.xlsx",
            data=data["excel_bytes"],
            file_name="assets_evaluation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    render_portfolio_history(data["portfolio_history"], data["history_skipped_assets"], data["timeline_events"])
    render_diagnostics(
        data["portfolio_history_raw"],
        data["portfolio_history"],
        data["portfolio_offsets"],
        data["snapshot_by_group"],
    )


if __name__ == "__main__":
    pd.options.mode.copy_on_write = True
    pd.options.future.infer_string = True
    main()
